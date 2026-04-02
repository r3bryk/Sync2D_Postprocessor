import pandas as pd
import os
import re
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.worksheet.dimensions import ColumnDimension
from typing import List, Tuple
from itertools import combinations
import matplotlib.pyplot as plt
import numpy as np
from sklearn.mixture import GaussianMixture
from sklearn.preprocessing import StandardScaler
from sklearn.preprocessing import normalize
import tkinter as tk
from tkinter.filedialog import askopenfilename
from scipy.stats import norm
from scipy.stats import gaussian_kde
from scipy.signal import argrelextrema
from scipy.spatial.distance import cosine
import warnings
warnings.filterwarnings("ignore", category=pd.errors.PerformanceWarning)



##########################################################################################
################################ Loading Input File ######################################
##########################################################################################

# Creating the file opening dialogue window
root = tk.Tk()
root.withdraw() # Prevents the tkinter window from coming up

# Opening the file
file_path = askopenfilename(title='Select CSV, TXT, or Excel file for processing...')
root.destroy()  # Closing the window

if not file_path:
    raise ValueError("WARNING: No file selected.", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

# Print start time and save it as start_time variable
start_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
print("")
print(150 * "*")
print(f"Processing started at: {start_time}")
print(150 * "-")

# Load the input file based on its extension
print(f"Loading the input file {file_path}...", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
if file_path.endswith('.xlsx'):
    df = pd.read_excel(file_path, engine='openpyxl')
elif file_path.endswith('.csv'):
    df = pd.read_csv(file_path, low_memory=False)
elif file_path.endswith('.txt'):
    df = pd.read_csv(file_path, delimiter='\t')
else:
    raise ValueError("Unsupported file format. Please select a CSV, TXT, or Excel file.")

print("File loaded successfully.", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
print(150 * "-")

# Check the first few rows
print("Input file info: ")
print(df.info())
print(150 * "-")
print("First 5 rows: ")
print(df.head())
print(150 * "-")



##########################################################################################
##################### Base Mass Retrieval and Feature Renaming ###########################
##########################################################################################

############################### Base Mass Retrieval ######################################
user_input_bm = input(
    "Would you like to retrieve base masses from spectra & calculate ratio Quant mass:Base mass?\n"
    "0 = No\n"
    "1 = Yes\n"
    "Type your choice and press Enter: "
).strip()
print(" ")

if user_input_bm == '0':
    print(f"Skipping base mass retrieval.")
    print(150 * "-")
    pass
else:
    # Base mass retrieval
    print("Retrieving base mass from spectrum & calculating ratio Quant mass:Base mass...", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

    # Check required columns and their data
    required_cols = ['Quant mass']
    spectrum_cols = ['Spectrum', 'Spectrum_Sync2D']
    existing_spectrum_col = next((col for col in spectrum_cols if col in df.columns and df[col].notna().any()), None)

    if not all(col in df.columns for col in required_cols) or existing_spectrum_col is None:
        print("WARNING: Missing one or more required columns (Quant mass or Spectrum/Spectrum_Sync2D) with data.")
    else:
        # Create 'Base mass' column if needed
        if 'Base mass' not in df.columns:
            df.insert(df.columns.get_loc('Quant mass') + 1, 'Base mass', np.nan)

        if df['Base mass'].notna().all():
            print("Base mass column already exists and is populated — skipping base mass retrieval.")
        else:
            def parse_spectrum(spectrum):
                if pd.isna(spectrum):
                    return []
                spectrum = str(spectrum).strip()
                if '(' in spectrum and '|' in spectrum:
                    # Format: (55|1000.00)(97|325.11)
                    pairs = spectrum.strip('()').split(')(')
                    return [(float(p.split('|')[0]), float(p.split('|')[1])) for p in pairs if '|' in p]
                else:
                    # Format: 55:1000.00 97:325.11
                    return [(float(p.split(':')[0]), float(p.split(':')[1])) for p in spectrum.split() if ':' in p]

            def get_base_mass(spectrum):
                mz_intensity = parse_spectrum(spectrum)
                if mz_intensity:
                    return max(mz_intensity, key=lambda x: x[1])[0]
                return np.nan

            df['Base mass'] = df[existing_spectrum_col].apply(get_base_mass)
            print("Base mass retrieval completed.", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

        # Skip Quant mass:Base mass ratio calculation if already calculated
        if 'Int(QM):Int(BM) (%)' in df.columns and df['Int(QM):Int(BM) (%)'].notna().any():
            print("Int(QM):Int(BM) (%) already exists and is populated — skipping QM/BM processing.")
            print(150 * "-")
        else:
            # Create 'Int(QM):Int(BM) (%)' column
            if 'Int(QM):Int(BM) (%)' not in df.columns:
                df.insert(df.columns.get_loc('Base mass') + 1, 'Int(QM):Int(BM) (%)', np.nan)

            def get_intensity_exact(spectrum, target_mass):
                if pd.isna(spectrum) or pd.isna(target_mass):
                    return np.nan
                try:
                    mz_intensity = parse_spectrum(spectrum)
                    matches = [intensity for mz, intensity in mz_intensity if int(round(mz)) == int(round(target_mass))]
                    return max(matches) if matches else np.nan
                except Exception:
                    return np.nan

            # Fix invalid values in 'Quant mass': replace NaN, 0, or non-numeric with Base mass
            df['Quant mass'] = pd.to_numeric(df['Quant mass'], errors='coerce')
            df['Quant mass'] = df['Quant mass'].where(df['Quant mass'] > 0, df['Base mass'])

            # Replace 'Quant mass' values not found in spectrum with Base mass
            def quant_mass_in_spectrum(row):
                q_mass = row['Quant mass']
                spec = row[existing_spectrum_col]
                mz_intensity = parse_spectrum(spec)
                return any(int(round(mz)) == int(round(q_mass)) for mz, _ in mz_intensity)

            df['Quant mass'] = df.apply(lambda row: row['Quant mass'] if quant_mass_in_spectrum(row) else row['Base mass'], axis=1)

            # Calculate intensities
            quant_intensities = df.apply(lambda row: get_intensity_exact(row[existing_spectrum_col], row['Quant mass']), axis=1)
            base_intensities = df.apply(lambda row: get_intensity_exact(row[existing_spectrum_col], row['Base mass']), axis=1)

            def compute_ratio(q_int, b_int):
                if pd.isna(q_int):
                    return 'WARNING: QM not found in spec'
                elif pd.notna(b_int) and b_int != 0:
                    ratio = (q_int / b_int) * 100
                    return round(ratio) if abs(ratio) >= 10 else ratio
                else:
                    return np.nan

            df['Int(QM):Int(BM) (%)'] = [
                compute_ratio(q, b) for q, b in zip(quant_intensities, base_intensities)
            ]

            print("Quant mass:Base mass ratio calculation completed.", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            print(150 * "-")



####################### Interactive Unknown Feature Renaming #############################
# Ask user whether to run interactive unknown feature renaming
user_input_rename = input(
    "Would you like to rename unknown features?\n"
    "0 = No\n"
    "1 = Yes\n"
    "Type your choice and press Enter: "
).strip()
print(" ")

if user_input_rename == "0":
    print("Skipping unknown features renaming.")
    print(150 * "-")
else:
    # Ask how unknown features are labeled in the 'Name' column
    print(
        "Choose the keyword that identifies unknown features in the 'Name' column:\n"
        "1 = Peak\n"
        "2 = Feature\n"
        "3 = Unknown\n"
        "4 = Custom keyword\n"
    )
    keyword_choice = input("Enter your choice (1, 2, 3, or 4): ").strip()
    if keyword_choice == "1":
        unknown_keyword = "Peak"
    elif keyword_choice == "2":
        unknown_keyword = "Feature"
    elif keyword_choice == "3":
        unknown_keyword = "Unknown"
    elif keyword_choice == "4":
        unknown_keyword = input("Type your custom keyword: ").strip()
    else:
        print("Invalid choice. Defaulting to 'Feature'.")
        unknown_keyword = "Feature"

    print(f"Unknown feature keyword set to: '{unknown_keyword}'\n")

    # Ask which columns define the renaming pattern
    print(
        "Specify the column to use as the FIRST identifier in the name.\n"
        "Common choices: R.I. calc, RT1, Retention index, Molecular ion, etc."
    )
    id_col1 = input("Enter column name for FIRST identifier: ").strip()
    print(" ")

    print(
        "Specify the column to use as the SECOND identifier in the name.\n"
        "Common choice: Med RT2 (sec), RT2, Retention index, Molecular ion, etc."
    )
    id_col2 = input("Enter column name for SECOND identifier: ").strip()
    print(" ")

    print(f"Identifier 1 column: {id_col1}")
    print(f"Identifier 2 column: {id_col2}\n")

    # Safety checks
    missing_cols = [c for c in [id_col1, id_col2] if c not in df.columns]
    if missing_cols:
        print("WARNING: The following identifier column(s) do not exist in the DataFrame: ")
        for c in missing_cols:
            print("  -", c)
        print("Unknown features renaming skipped.")
        print(150 * "!")
    else:
        # Create renaming function
        print("Renaming unknown features... ", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        def rename_unknown_feature(row):
            name = row["Name"]

            # Only rename if 1) it's a string AND 2) contains the unknown keyword
            if isinstance(name, str) and unknown_keyword.lower() in name.lower():
                v1 = row[id_col1]
                v2 = row[id_col2]

                # Convert identifiers (numeric values) safely
                # First identifier: rounded integer
                try:
                    ident1 = str(int(round(float(v1))))
                except:
                    ident1 = "NA"

                # Second identifier: keep RT2 3 decimals (string)
                try:
                    ident2 = f"{round(float(v2), 3):.3f}"
                except:
                    ident2 = "NA"

                # Format: Keyword_ident1_ident2
                return f"{unknown_keyword}_{ident1}_{ident2}"

            # Otherwise return the original name
            return name

        # Apply renaming
        df["Name"] = df.apply(rename_unknown_feature, axis=1)

        print("Feature renaming completed. ", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        print(150 * "-")
        # Print first 10 entries for QC
        print("First 10 entries after renaming: ")
        print(df[['Name', id_col1, id_col2]].head(10))
        print(150 * "-")



################# Move 'Med RT1' and 'Med RT2' before 'R.I. calc' ########################
for rt_col in ["Med RT1 (sec)", "Med RT2 (sec)"]:
    if rt_col in df.columns:
        col_data = df.pop(rt_col)
        df.insert(df.columns.get_loc("R.I. calc"), rt_col, col_data)



##########################################################################################
################################# Cutoff Estimation ######################################
##########################################################################################

##################### Identify Metadata and Sample Peak Area Columns #####################
meta_cols = ['ID', 'Name', 'Formula', 'M.W.', 'Quant mass', 'R.I. calc', 'S/N', 'Samples']
# sample_start_idx = df.columns.get_loc("Samples") + 1
# Ask user to specify the column name after which sample peak area columns start
col_name = input("Enter the name of the column after which sample peak area columns start: ").strip()

# Validate that the column exists in df
if col_name not in df.columns:
    raise KeyError(f"WARNING: Column '{col_name}' not found in DataFrame. Please check spelling and try again.\n")

# Get the index of the column after which sample peak area columns start
sample_start_idx = df.columns.get_loc(col_name) + 1
print(f"Sample peak area columns will be taken starting from index {sample_start_idx} (after '{col_name}').\n")

sample_cols = df.columns[sample_start_idx:]
area_columns = df.columns[sample_start_idx:]



################################### Cutoff Estimation ####################################
# Ask user whether to run cutoff estimation
choice = input(
    "Would you like to perform real vs. pseudo peak cutoff estimation?\n"
    "0 = No\n"
    "1 = Yes\n"
    "Type your choice and press Enter: "
).strip()
print(" ")

if choice not in ["0", "1"]:
    print("WARNING: Invalid input — skipping cutoff estimation.")
    print(150 * "-")
    run_cutoff = False
elif choice == "0":
    print("Cutoff estimation skipped by user.")
    print(150 * "-")
    run_cutoff = False
else:
    run_cutoff = True

# Skip entire section and continue script
if not run_cutoff:
    pass
else:
    ################################# Histogram Analysis #####################################
    # Flatten all area values into a single Series for histogram analysis
    area_values = df[area_columns].values.flatten()

    # Drop NaNs and zeroes (since zeros likely represent no peak)
    area_values = area_values[~pd.isna(area_values)]
    area_values = area_values[area_values > 0]

    # Describe the basic statistics
    area_stats = pd.Series(area_values).describe(percentiles=[.25, .5, .75, .9, .95, .99])
    print(150 * "-")
    print("Area columns stats: ")
    print(area_stats)
    print(150 * "-")

    print("Plotting histogram — see new window... ")
    print(150 * "-")
    # Plot histogram on a log scale to reveal distribution
    plt.figure(figsize=(14, 8))
    log_area_values = np.log10(area_values)
    plt.hist(log_area_values, bins=100, density=True, alpha=0.5, color='lightgray', edgecolor='black')
    plt.xlabel('Log10(Peak Area)')
    plt.ylabel('Density')
    plt.title('Distribution of Log10(Peak Area) across all samples')
    plt.grid(True)
    plt.tight_layout()
    plt.show()



    ########################### Kernel Density Estimation (KDE) ##############################
    print(f"Kernel Density Estimation (KDE) in progress...\n")
    # Log10-transform and filter the area values
    log_area = np.log10(area_values)
    # Estimate density using Kernel Density Estimation (KDE)
    kde = gaussian_kde(log_area, bw_method=0.2)  # Tune bw_method (e.g., 0.1 to 0.3) if needed
    x_vals = np.linspace(log_area.min(), log_area.max(), 1000)
    kde_vals = kde(x_vals)

    # Find local minima in the KDE
    local_minima_idx = argrelextrema(kde_vals, np.less)[0]

    # If found, select the first minimum between the two KDE peaks
    if len(local_minima_idx) > 0:
        cutoff_kde_log10 = x_vals[local_minima_idx[0]]
        cutoff_kde = 10 ** cutoff_kde_log10
        print(f"KDE-based log10(area) cutoff: {cutoff_kde_log10:.2f}")
        print(f"KDE-based area cutoff: {cutoff_kde:,.0f}")
        print(150 * "-")
    else:
        cutoff_kde_log10 = None
        cutoff_kde = None
        print("!" * 150)
        print("WARNING: No local minima found in KDE curve.")
        print("!" * 150)



    ############################ Gaussian Mixture Model (GMM) ################################
    print(f"Gaussian Mixture Modeling (GMM) in progress...\n")
    # Reshape for sklearn and fit GMM with 2 components (pseudo vs real peaks)
    log_area = np.log10(area_values).reshape(-1, 1)
    gmm = GaussianMixture(n_components=2, random_state=42)
    gmm.fit(log_area)

    # Get means and stds of components
    means = gmm.means_.flatten()
    stds = np.sqrt(gmm.covariances_.flatten())

    # Sort components to get the lower (pseudo) and higher (real) mean
    low_idx, high_idx = np.argsort(means)

    # Approximate intersection of the Gaussians
    # Formula for two Gaussians: solve N1(x) = N2(x)
    mu1, mu2 = means[low_idx], means[high_idx]
    sigma1, sigma2 = stds[low_idx], stds[high_idx]
    cutoff_gmm_log10 = (mu1 * sigma2**2 - mu2 * sigma1**2 + sigma1 * sigma2 * np.sqrt((mu1 - mu2)**2 + 2 * (sigma2**2 - sigma1**2) * np.log(sigma2 / sigma1))) / (sigma2**2 - sigma1**2)
    cutoff_gmm = 10 ** cutoff_gmm_log10

    print(f"GMM-based log10(area) cutoff: {cutoff_gmm_log10:.2f}")
    print(f"GMM-based area cutoff: {cutoff_gmm:,.0f}")
    print(150 * "-")

    # Generate x values over log10(area) range
    x = np.linspace(log_area.min(), log_area.max(), 1000).flatten()

    # Compute individual Gaussian PDFs
    pdf1 = norm.pdf(x, mu1, sigma1)
    pdf2 = norm.pdf(x, mu2, sigma2)



    ########################### GMM & KDE Cutoffs & Mean Cutoff ##############################
    # Get 3 cutoff values between the GMM and KDE cutoffs: cutoff_high (more conservative (higher)) value),
    # cutoff_low, and cutoff_mean (can be used instead of cutoff_high for further calculations)
    # cutoff_high is typically safer for filtering out pseudo peaks
    # Handle cases where cutoff_kde is None as no local minima was found in KDE curve 
    if cutoff_kde is None:
        cutoff_high = cutoff_gmm
        cutoff_low = cutoff_gmm
        cutoff_mean = cutoff_gmm
        cutoff_kde_str = "N/A"
    else:
        cutoff_high = max(cutoff_gmm, cutoff_kde)
        cutoff_low = min(cutoff_gmm, cutoff_kde)
        cutoff_mean = (cutoff_gmm + cutoff_kde) / 2
        cutoff_kde_str = f"{cutoff_kde:,.0f}"

    cutoff_mean_log10 = np.log10(cutoff_mean)

    print(f"Peak cutoff (higher): {cutoff_high:,.0f}")
    print(f"Peak cutoff (lower): {cutoff_low:,.0f}")
    print(f"Peak cutoff (mean): {cutoff_mean:,.0f}")
    print(150 * "-")
    
    print("Plotting graph — see new window... ")
    print(150 * "-")
    # Plot GMM, KDE, GMM cutoff, KDE cutoff, and mean cutoff
    plt.figure(figsize=(14, 8))
    plt.hist(log_area, bins=100, density=True, alpha=0.5, color='lightgray', label='Histogram (Log10(Peak Area))')
    # GMM curves and cutoff
    plt.plot(x, pdf1, label=f'Pseudo Peak GMM (Mean: {mu1:.2f})', color='lightgreen', linestyle='-')
    plt.plot(x, pdf2, label=f'Real Peak GMM (Mean: {mu2:.2f})', color='darkgreen', linestyle='-')
    plt.axvline(cutoff_gmm_log10, color='green', linestyle='--', label=f'GMM Cutoff (Log10): {cutoff_gmm_log10:.2f}\nGMM Cutoff: {cutoff_gmm:,.0f}')
    # KDE curve and cutoff
    plt.plot(x_vals, kde_vals, color='orange', lw=2, label='KDE Curve')
    if cutoff_kde is not None and cutoff_kde_log10 is not None:
        plt.axvline(cutoff_kde_log10, color='darkorange', linestyle='--', label=f'KDE Cutoff (Log10): {cutoff_kde_log10:.2f}\nKDE Cutoff: {cutoff_kde:,.0f}')
    # Mean cutoff
    plt.axvline(cutoff_mean_log10, color='red', linestyle='--', label=f'Mean Cutoff (Log10): {cutoff_mean_log10:.2f}\nMean Cutoff: {cutoff_mean:,.0f}')
    # Labels and legend
    plt.title('GMM and KDE Fits of Log10(Peak Area)')
    plt.xlabel('Log10(Peak Area)')
    plt.ylabel('Density')
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    plt.show()



##########################################################################################
######################### Convert Spectra and Merge Features #############################
##########################################################################################

################################### Convert Spectra ######################################
def sync2d_spec_converter(df):
    """
    Rename and transform the 'Spectrum' column containing Sync2D-style values.
    Steps:
    - If 'Spectrum_Sync2D' and 'Spectrum' already exist, skip conversion
    Otherwise:
    - Move 'Spectrum' before specified column and rename it to 'Spectrum_Sync2D'
    - Create an empty 'Spectrum' column and place it after 'Spectrum_Sync2D'
    - Convert format from '(77|871.48)(156|1000.00)' to '77:871.48 156:1000.00'
    """
    print("Converting spectra...", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

    if 'Spectrum_Sync2D' in df.columns and 'Spectrum' in df.columns:
        print("Both 'Spectrum_Sync2D' and 'Spectrum' columns already exist — skipping conversion.")
        print(150 * "-")
        return df
    
    # Step 1: Locate 'Spectrum' column
    if 'Spectrum' not in df.columns:
        raise ValueError("WARNING: Required column 'Spectrum' not found in the input DataFrame.")

    spectrum_idx = df.columns.get_loc('Spectrum')
    # Ask user to specify the column before which 'Spectrum' should be inserted
    col_name = input("Enter the name of the column before which 'Spectrum' should be inserted: ").strip()

    # Validate that the column exists in df
    if col_name not in df.columns:
        raise KeyError(f"WARNING: Column '{col_name}' not found in DataFrame. Please check spelling and try again.\n")

    # Get the index of the column before which to insert 'Spectrum'
    samples_idx = df.columns.get_loc(col_name)
    print(f"'Spectrum' will be inserted in index 22 before column '{col_name}'.")

    # Step 2: Rename 'Spectrum' -> 'Spectrum_Sync2D'
    df.rename(columns={'Spectrum': 'Spectrum_Sync2D'}, inplace=True)

    # Step 3: Insert new empty 'Spectrum' column right after the renamed one
    df.insert(samples_idx, 'Spectrum', '')  # This places new 'Spectrum' before the specified column 

    # Step 4: Convert spectrum string format
    def convert_spectrum_format(value):
        if pd.isna(value):
            return ''
        # Replace e.g., (77|871.48) -> 77:871.48
        return ' '.join(f"{m}:{i}" for m, i in re.findall(r"\((\d+)\|([\d.]+)\)", value))

    df['Spectrum'] = df['Spectrum_Sync2D'].apply(convert_spectrum_format)

    print("Spectra conversion completed.", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    print(150 * "-")
    return df

# Use spectrum converter
df = sync2d_spec_converter(df)



################################### Merge Features #######################################
def parse_spectrum(spectrum_str: str) -> dict:
    """Parses a spectrum string like '77:871.48 156:1000.00' into a dict."""
    if pd.isna(spectrum_str) or not isinstance(spectrum_str, str):
        return {}
    return {int(p.split(':')[0]): float(p.split(':')[1]) for p in spectrum_str.split() if ':' in p}


def spectrum_similarity(s1: dict, s2: dict, mz_tol: float = 0.5, method: str = "DISCO") -> float:
    """Computes similarity between two spectra using DISCO or NIST NDP."""
    all_mz = sorted(set(s1) | set(s2))
    v1 = np.array([s1.get(mz, 0) for mz in all_mz])
    v2 = np.array([s2.get(mz, 0) for mz in all_mz])
    if method.upper() == "NDP":
        sim = np.dot(v1, v2) / (np.linalg.norm(v1) * np.linalg.norm(v2) + 1e-10)
    else:  # DISCO (mean-centered cosine)
        v1 -= v1.mean()
        v2 -= v2.mean()
        sim = 1 - cosine(v1, v2)
    return sim


def merge_internal_features(df: pd.DataFrame,
                            input_file: str,
                            rt1_tol: Tuple[float, float],
                            rt2_tol: Tuple[float, float],
                            unknown_keyword: str,
                            sim_thresh: float = 0.7,
                            mz_tol: float = 0.1,
                            similarity_method: str = "DISCO") -> pd.DataFrame:
    """
    Merge unknown features based on RT and spectral similarity.
    Saves merged unknowns and the final merged features to separate Excel files.
    """

    user_in = input(
        "Would you like to merge internal features?\n"
        "1 = Yes (perform merging)\n"
        "0 = No (skip merging)\n"
        "Type your choice and press Enter: "
    ).strip()
    print(" ")

    if user_in == "0":
        print("Skipping internal feature merging.")
        print(150 * "-")
        return df.copy()  # Return original df unchanged
    
    # Continue with merging only if user_in == "1"
    print("Merging unknown features...", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

    df_all = df.copy()

    # Ensure 'Merged features' column exists after 'ID'
    if 'Merged features' not in df_all.columns:
        id_index = df_all.columns.get_loc('ID') + 1
        df_all.insert(id_index, 'Merged features', '')

    # Check if unknown_keyword exists, if not, use 'Feature'
    if not unknown_keyword:
        print("WARNING: 'unknown_keyword' not provided — defaulting to 'Feature'.")
        unknown_keyword = "Feature"
        
    # Split knowns and unknowns
    df_unknowns = df_all[df_all['Name'].str.contains(unknown_keyword, case=False, na=False)].copy()
    df_knowns = df_all[~df_all.index.isin(df_unknowns.index)].copy()
    print(f"Found {len(df_unknowns)} unknown and {len(df_knowns)} known features.\n")

    if df_unknowns.empty:
        print("WARNING: No unknown features found — skipping merging.")
        return df_all.copy()

    # Calculate total area across all sample peak area columns
    # Ask user to specify the column name after which sample peak area columns start
    unk_col_name = input("Enter the name of the column after which sample peak area columns start: ").strip()

    # Validate that the column exists in df
    if unk_col_name not in df_unknowns.columns:
        raise KeyError(f"WARNING: Column '{unk_col_name}' not found in DataFrame. Please check spelling and try again.\n")

    area_cols = df_unknowns.columns[df_unknowns.columns.get_loc(unk_col_name) + 1:]
    print(f"Sample peak area columns will be taken after '{unk_col_name}'.\n")

    df_unknowns['TotalArea'] = df_unknowns[area_cols].sum(axis=1)

    rt1 = df_unknowns['Med RT1 (sec)'].values
    rt2 = df_unknowns['Med RT2 (sec)'].values
    base_mass = df_unknowns['Base mass'].values

    used = np.full(len(df_unknowns), False)
    merged_log = []
    merged_main_indices = []  # Store index labels (not positions)

    for i in range(len(df_unknowns)):
        if used[i]:
            continue

        try:
            main_spec = parse_spectrum(df_unknowns.iloc[i]['Spectrum'])
        except Exception:
            print(f"Skipping feature {df_unknowns.iloc[i]['ID']} due to spectrum parsing error.")
            continue

        group = [i]
        candidates = 0
        matched = 0

        for j in range(i + 1, len(df_unknowns)):
            if used[j]:
                continue

            # RT window conditions
            if (rt1[j] >= rt1[i] - rt1_tol[0] and rt1[j] <= rt1[i] + rt1_tol[1] and
                rt2[j] >= rt2[i] - rt2_tol[0] and rt2[j] <= rt2[i] + rt2_tol[1]):

                # OPTIONAL: Match on Base Mass
                # Uncomment the following block (4 lines) to require Base mass to match for merging.
                # Comment the 4 lines if Base mass matching is not required.
                if pd.isna(base_mass[i]) or pd.isna(base_mass[j]):
                    continue
                if base_mass[i] != base_mass[j]:
                    continue

                candidates += 1
                try:
                    other_spec = parse_spectrum(df_unknowns.iloc[j]['Spectrum'])
                    sim1 = spectrum_similarity(main_spec, other_spec, mz_tol, similarity_method)
                    sim2 = spectrum_similarity(other_spec, main_spec, mz_tol, similarity_method)
                    sim_final = min(sim1, sim2)
                except Exception:
                    print(f"Failed similarity comparison between {df_unknowns.iloc[i]['ID']} and {df_unknowns.iloc[j]['ID']}")
                    continue

                if sim_final >= sim_thresh:
                    group.append(j)
                    used[j] = True
                    matched += 1

        if len(group) > 1:
            print(f"Feature {df_unknowns.iloc[i]['ID']} grouped with {matched} of {candidates} RT neighbors.")

            group_df = df_unknowns.iloc[group]
            main_pos = group_df['TotalArea'].idxmax()
            main_idx_label = df_unknowns.loc[main_pos].name  # Index label from original df

            group_indices_labels = df_unknowns.iloc[group].index.tolist()
            group_indices_sorted = sorted(group_indices_labels, key=lambda x: df.loc[x, area_cols].sum(), reverse=True)
            merged_str = "_".join(str(df.loc[idx, 'ID']) for idx in group_indices_sorted)

            # Sum area columns into main
            for col in area_cols:
                df.loc[main_idx_label, col] = df.loc[group_indices_labels, col].sum()

            df.loc[main_idx_label, 'Merged features'] = merged_str

            # Drop all other rows in group
            drop_indices = [idx for idx in group_indices_labels if idx != main_idx_label]
            df = df.drop(index=drop_indices)

            merged_log.append(group_indices_sorted)
            merged_main_indices.append(main_idx_label)

    print(" ")
    print(f"Merged groups: {len(merged_log)}")
    print(f"Retained unknowns after merging: {len(merged_main_indices)}")

    if not merged_log:
        print("WARNING: No merging performed. Returning original data.")
        print(150 * "-")
        return df_all.copy()

    # Identify unmerged unknowns
    merged_indices_flat = set(sum(merged_log, []))  # This needs to be converted to a list
    unmerged_indices = [idx for idx in df_unknowns.index if idx not in merged_indices_flat]
    print(f"Retained unmerged unknowns: {len(unmerged_indices)}")

    # Recombine final data (include knowns, merged mains, and unmerged unknowns)
    print(f"Saving files with merged & unmerged unknowns...")
    final_df = pd.concat([df_knowns,
                          df.loc[merged_main_indices],
                          df.loc[unmerged_indices]], ignore_index=True)

    # Save merged unknowns for inspection
    merged_df_for_export = df_all.loc[list(merged_indices_flat)].copy()  # Convert set to list here
    merged_df_for_export['ID_Mrg'] = np.nan

    # Move 'ID_Mrg' after 'ID'
    cols = list(merged_df_for_export.columns)
    if 'ID_Mrg' in cols:
        cols.insert(cols.index('ID') + 1, cols.pop(cols.index('ID_Mrg')))
        merged_df_for_export = merged_df_for_export[cols]

    for idx, group in enumerate(merged_log, start=1):
        for i in group:
            merged_df_for_export.at[i, 'ID_Mrg'] = idx

    merged_df_for_export.to_excel(input_file.replace(".xlsx", "_MrgdFtrs.xlsx"), index=False)
    merged_unknowns_only = final_df[final_df['Name'].str.contains(unknown_keyword, case=False, na=False)].copy()
    merged_unknowns_only.to_excel(input_file.replace(".xlsx", "_Mrgd&Unmrgd.xlsx"), index=False)

    print("Merging completed.", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    print(150 * "-")
    return final_df

# Merge internal features
merged_df = merge_internal_features(
    df=df, # Input df
    input_file=file_path, # Input df path
    rt1_tol=(10, 10), # RT1 left & right tolerance values (sec)
    rt2_tol=(0.5, 0.5), # RT2 bottom & top tolerance values (sec)
    unknown_keyword=unknown_keyword, # Unknown feature keyword, e.g., Feature, Peak, Unknown, or custom
    sim_thresh=0.7, # Spectral similarity threshold
    mz_tol=0.1, # m/z tolerance
    similarity_method="NDP"  # Spectral similarity method: "DISCO" (DIstance & Spectrum Correlation Optimization) or "NDP" (NIST Normalized Dot Product)
)
df = merged_df # Pass df with merged unknowns for further processing



##########################################################################################
########################## Detection Frequency Calculation ###############################
##########################################################################################

# Ask user whether to run LOD specification
choice = input(
    "Would you like to perform LOD specification (cutoff value)?\n"
    "0 = No\n"
    "1 = Yes\n"
    "Type your choice and press Enter: "
).strip()
print(" ")

if choice not in ["0", "1"]:
    print("WARNING: Invalid input — skipping LOD specification.")
    print(150 * "-")
    run_lod = False
elif choice == "0":
    print("LOD specification skipped by user.")
    print(150 * "-")
    run_lod = False
else:
    run_lod = True

# Skip entire section and continue script
if not run_lod:
    pass
else:
    # Interactive cutoff value selection
    print("Selecting cutoff value to use for pseudo peak replacement & LOD... ")
    # Determine whether cutoff values exist
    have_low  = 'cutoff_low'  in globals() and int(cutoff_low)  is not None
    have_high = 'cutoff_high' in globals() and int(cutoff_high) is not None
    have_mean = 'cutoff_mean' in globals() and int(cutoff_mean) is not None

    # If no cutoff values exist, force manual input
    if not (have_low or have_high or have_mean):
        print("WARNING: No estimated cutoff values found. You must enter a custom cutoff value.")
        
        while True:
            user_cut = input("Enter custom cutoff value (positive number, e.g., 18250): ").strip()
            try:
                cutoff = float(user_cut)
                if cutoff > 0:
                    break
            except:
                pass
            print("Invalid input. Please enter a positive numeric value.")
        
        print(f"Using custom cutoff: {int(cutoff)}")
        print(150 * "-")
    else:
        # Display available options (only those that exist)
        option_num = 1
        option_map = {}

        if have_low:
            print(f"{option_num} - cutoff_low     = {cutoff_low}")
            option_map[str(option_num)] = ('low', cutoff_low)
            option_num += 1

        if have_high:
            print(f"{option_num} - cutoff_high    = {cutoff_high}")
            option_map[str(option_num)] = ('high', cutoff_high)
            option_num += 1

        if have_mean:
            print(f"{option_num} - cutoff_mean    = {cutoff_mean}")
            option_map[str(option_num)] = ('mean', cutoff_mean)
            option_num += 1

        print(f"{option_num} - Enter custom value")
        option_map[str(option_num)] = ('custom', None)

        print(150 * "-")

        # Ask user for choice
        while True:
            choice = input("Type your choice: ").strip()
            if choice in option_map:
                break
            print("Invalid choice. Try again.")

        choice_type, choice_value = option_map[choice]

        if choice_type == 'custom':
            # Manual custom value input
            while True:
                user_cut = input("Enter custom cutoff value (positive number, e.g., 18250): ").strip()
                try:
                    cutoff = float(user_cut)
                    if cutoff > 0:
                        break
                except:
                    pass
                print("Invalid input. Please enter a positive numeric value.")
        else:
            # Use the chosen estimated cutoff
            cutoff = float(choice_value)

        print(150 * "-")
        print("Selected pseudo peak cutoff & LOD:")
        print(int(cutoff))
        print(150 * "-")
        # cutoff = 18250  # Based on the modeling for the whole initial dataset (ca. 40,000 features)

    # Use cutoff as LOD
    # Ask user to specify the column name before which to insert 'LOD' column
    col_name = input("Enter the name of the column before which to insert 'LOD' column: ").strip()

    # Validate that the column exists in df
    if col_name not in df.columns:
        raise KeyError(f"WARNING: Column '{col_name}' not found in DataFrame. Please check spelling and try again.\n")

    # Get the index of the column after which sample peak area columns start
    smpl_idx = df.columns.get_loc(col_name)
    print(f"'LOD' column will be inserted before '{col_name}'.")

    # Only create and fill 'LOD' column if it does NOT exist,
    # OR exists but is completely empty (all NaN / empty / None)
    if ('LOD' not in df.columns) or (df['LOD'].isna().all()):
        # Insert column only if it does not exist
        if 'LOD' not in df.columns:
            df.insert(smpl_idx, 'LOD', np.nan)  # Place new 'LOD' column before 'Samples'
        # Fill 'LOD' column with cutoff values
        df['LOD'] = cutoff
        print(f"'LOD' column populated with cutoff = {int(cutoff)}")
        print(150 * "-")
    else:
        print("'LOD' column already exists and is populated — skipping creation.")
        print(150 * "-")


# Define blank columns
umu_blank_cols = ['Blank_01_241112_012', 'Blank_02_241113_026', 'Blank_03_241114_044', 'Blank_04_241115_064']
all_blank_cols = ['Blank_01_241112_012', 'Blank_02_241113_026', 'Blank_03_241114_044', 'Blank_04_241115_064',
                  'CZ_FB_01_241112_017', 'CZ_FB_02_241115_058', 'CZ_FB_03_241113_027', 'EE_FB_01_241112_022',
                  'EE_FB_02_241112_021', 'EE_FB_03_241114_048', 'IT_FB_02_241113_031', 'NL_FB_01_241115_052',
                  'NL_FB_02_241113_035', 'NL_FB_03_241113_032', 'PT_FB_01_241115_060', 'PT_FB_02_241115_059',
                  'PT_FB_03_241115_055', 'SI_FB_01_241112_018', 'SI_FB_02_241112_014', 'SI_FB_03_241112_016',
                  'UK_FB_01_241113_028', 'UK_FB_02_241112_020']
field_blank_cols = ['CZ_FB_01_241112_017', 'CZ_FB_02_241115_058', 'CZ_FB_03_241113_027', 'EE_FB_01_241112_022',
                  'EE_FB_02_241112_021', 'EE_FB_03_241114_048', 'IT_FB_02_241113_031', 'NL_FB_01_241115_052',
                  'NL_FB_02_241113_035', 'NL_FB_03_241113_032', 'PT_FB_01_241115_060', 'PT_FB_02_241115_059',
                  'PT_FB_03_241115_055', 'SI_FB_01_241112_018', 'SI_FB_02_241112_014', 'SI_FB_03_241112_016',
                  'UK_FB_01_241113_028', 'UK_FB_02_241112_020']


# Ask user whether to run detection frequency calculations
choice = input(
    "Would you like to perform detection frequency (DF) calculations?\n"
    "0 = No\n"
    "1 = Yes\n"
    "Type your choice and press Enter: "
).strip()
print(" ")

if choice not in ["0", "1"]:
    print("WARNING: Invalid input — skipping detection frequency (DF) calculations.")
    print(150 * "-")
    run_df = False
elif choice == "0":
    print("Detection frequency calculations (DF) skipped by user.")
    print(150 * "-")
    run_df = False
else:
    run_df = True

# Skip entire section and continue script
if not run_df:
    pass
else:
    ######################### Detection Frequency (DF) Calculation ###########################
    print("Starting detection frequency (DF) calculations...", datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "\n")

    # List of DN/DF columns to check for existence
    dn_df_cols = ['DN', 'DF (%)', 'DN blanks', 'DF blanks (%)', 'DN UMU blanks', 'DF UMU blanks (%)']
    # Check if all columns exist and have non-null data
    cols_exist_and_filled = all(col in df.columns and df[col].notna().any() for col in dn_df_cols)

    if not cols_exist_and_filled:
        # Ask user to specify the column name after which real sample peak area columns start
        col_name = input(
            "Enter the name of the column after which real sample peak area columns start. \n"
            "NB! Think of excluding QC & reference standard samples. \n"
        ).strip()

        # Validate that the column exists in df
        if col_name not in df.columns:
            raise KeyError(f"WARNING: Column '{col_name}' not found in DataFrame. Please check spelling and try again.\n")

        # Get the index of the column after which real sample peak area columns start
        start_col = df.columns.get_loc(col_name) + 1

        print(f"Real sample peak area columns will be taken starting from index {start_col} (after '{col_name}').\n")
        sample_cols_for_df = df.columns[start_col:]

        # Exclude 'FB' and UMU blank columns
        real_sample_cols = [
            col for col in sample_cols_for_df
            if 'FB' not in col and col not in umu_blank_cols
        ]
        N_real_samples = len(real_sample_cols)
        print(f"N_real_samples: {N_real_samples}. \nShould be 174 samples. \nNB! There are 18 FBs & 4 UMU blanks that are excluded from real samples\n")

        # Ensure numeric types for detection frequency calculation
        df[real_sample_cols] = df[real_sample_cols].apply(pd.to_numeric, errors='coerce')
        df[all_blank_cols] = df[all_blank_cols].apply(pd.to_numeric, errors='coerce')
        df[umu_blank_cols] = df[umu_blank_cols].apply(pd.to_numeric, errors='coerce')

        # Calculate detection numbers (DN)
        df["DN"] = (df[real_sample_cols] > cutoff).sum(axis=1)
        df["DN blanks"] = (df[all_blank_cols] > cutoff).sum(axis=1)
        df["DN UMU blanks"] = (df[umu_blank_cols] > cutoff).sum(axis=1)
        print(f"Calculated DN, DN blanks, DN UMU blanks.\n")

        # Calculate detection frequencies (DF)
        df["DF (%)"] = (df["DN"] / N_real_samples * 100).round(0)
        df["DF blanks (%)"] = (df["DN blanks"] / len(all_blank_cols) * 100).round(0)
        df["DF UMU blanks (%)"] = (df["DN UMU blanks"] / len(umu_blank_cols) * 100).round(0)
        print(f"Calculated DF (%), DF blanks (%), DF UMU blanks (%).\n")

        # Move DN and DF columns before specified column
        # Ask user to specify the column name before which DN and DF columns should be inserted
        col_name = input("Enter the name of the column before which DN and DF columns should be inserted: ").strip()

        # Validate that the column exists in df
        if col_name not in df.columns:
            raise KeyError(f"WARNING: Column '{col_name}' not found in DataFrame. Please check spelling and try again.\n")

        # Get the index of the column before which DN and DF columns should be inserted
        start_col = df.columns.get_loc(col_name)
        print(f"DN and DF columns will be inserted before '{col_name}'.\n")
        
        for col in ['DN', 'DF (%)', 'DN blanks', 'DF blanks (%)', 'DN UMU blanks', 'DF UMU blanks (%)']:
            if col in df.columns:
                col_data = df.pop(col)
                df.insert(df.columns.get_loc(col_name), col, col_data)

        # Country-specific DN/DF calculation
        country_codes = ['CZ', 'EE', 'IT', 'NL', 'PT', 'SI', 'UK']
        for country in country_codes:
            country_cols = [col for col in real_sample_cols if col.startswith(country)]
            df[f'DN {country}'] = (df[country_cols] > cutoff).sum(axis=1)
            df[f'DF {country} (%)'] = (df[f'DN {country}'] / len(country_cols) * 100).round(0)
            print(f"Total real samples in {country}: {len(country_cols)}")
        print(150 * "-")

        # Reorder country-specific DN/DF columns
        country_order = ['CZ', 'EE', 'IT', 'NL', 'PT', 'SI', 'UK']
        for country in (country_order):
            df.insert(df.columns.get_loc(col_name), f'DF {country} (%)', df.pop(f'DF {country} (%)'))
            df.insert(df.columns.get_loc(f'DF {country} (%)'), f'DN {country}', df.pop(f'DN {country}'))
    else:
        print("DN/DF columns already exist and are populated — skipping calculation.")
        print(150 * "-")
    

# Ask user whether to run DF- & blank-based feature prioritization
choice = input(
    "Would you like to perform DF-based feature prioritization?\n"
    "0 = No\n"
    "1 = Yes\n"
    "Type your choice and press Enter: "
).strip()
print(" ")

if choice not in ["0", "1"]:
    print("WARNING: Invalid input — skipping DF-based feature prioritization.")
    print(150 * "-")
    run_prio = False
elif choice == "0":
    print("DF-based feature prioritization skipped by user.")
    print(150 * "-")
    run_prio = False
else:
    run_prio = True

# Skip entire section and continue script
if not run_prio:
    pass
else:
    ############################### DF-based Prioritization ##################################
    print("Starting DF- & blank-based feature prioritization...", datetime.now().strftime('%Y-%m-%d %H:%M:%S'), "\n")
    
    # List of columns created/used in DF-based prioritization
    priority_cols = [
        'Priority', 'Reason', 'Ranking'
    ]

    # Check if all priority columns exist and are populated (at least one non-null value)
    priority_cols_exist_and_filled = all(col in df.columns and df[col].notna().any() for col in priority_cols)

    if not priority_cols_exist_and_filled:
        # Define exemptions
        exempt = (
            df["Keep"].isin(["IS", "RS", "Mix1", "Mix2", "Mix1&2", "X", "(X)"]) |
            df["CL Sync2D"].isin([1, 2]) |
            df["TargetHit"].notna() |
            df["NormanHit"].notna()
        )
        exempt = exempt.fillna(False)  # Ensure no NaN issues

        # Insert DN/DF columns for unknowns after 'DF UMU blanks (%)'
        df.insert(df.columns.get_loc('DF UMU blanks (%)') + 1, 'DN unknowns', np.nan)
        df.insert(df.columns.get_loc('DF UMU blanks (%)') + 2, 'DF unknowns (%)', np.nan)

        # Insert 'Priority' column before specified column
        # Ask user to specify the column name before which to insert 'Priority' column 
        col_name = input("Enter the name of the column before which to insert 'Priority' column: ").strip()

        # Validate that the column exists in df
        if col_name not in df.columns:
            raise KeyError(f"WARNING: Column '{col_name}' not found in DataFrame. Please check spelling and try again.\n")

        # Get the index of the column after which sample peak area columns start
        sample_idx = df.columns.get_loc(col_name)
        print(f"'Priority' column will be inserted before '{col_name}'.\n")
        
        df.insert(sample_idx, 'Priority', np.nan)
        # Insert step-specific priority columns
        df.insert(sample_idx, 'Priority_4', np.nan)
        df.insert(sample_idx, 'Priority_3', np.nan)
        df.insert(sample_idx, 'Priority_2', np.nan)
        df.insert(sample_idx, 'Priority_1', np.nan)
        df.insert(df.columns.get_loc('Priority') + 1, 'Reason', "")
        df.insert(df.columns.get_loc('Reason') + 1, 'Ranking', np.nan)

        # Insert 'Area_Mean(UMU blanks)' column after 'DF unknowns (%)'
        df.insert(df.columns.get_loc('DF UK (%)') + 1, 'Area_Mean(UMU blanks)', np.nan)

        # Calculate Area_Mean(UMU blanks)
        df['Area_Mean(UMU blanks)'] = df[umu_blank_cols].apply(lambda row: row[row > 0].mean(), axis=1)
        df['Area_Mean(UMU blanks)'] = df['Area_Mean(UMU blanks)'].apply(lambda x: np.ceil(x) if pd.notna(x) else np.nan)

        # Insert 'Area_Mean(Samples)' before 'Area_Mean(UMU blanks)'
        col_index = df.columns.get_loc('Area_Mean(UMU blanks)')
        df.insert(col_index, 'Area_Mean(Samples)', np.nan)

        # Calculate mean area for real samples with areas > cutoff
        df['Area_Mean(Samples)'] = df[real_sample_cols].apply(
            lambda row: row[row > cutoff].mean(), axis=1
        )

        # Insert new column 'Area_Mean_Ratio(UMU)' after 'Area_Mean(Samples)'
        col_index = df.columns.get_loc('Area_Mean(UMU blanks)') + 1
        df.insert(col_index, 'Area_Mean_Ratio(UMU)', np.nan)

        # Calculate ratio and round to integer
        df['Area_Mean_Ratio(UMU)'] = (
            df['Area_Mean(Samples)'] / df['Area_Mean(UMU blanks)']
        ).round(0).astype('Int64')

        # Insert 'Area_Median_Ratio(UMU)' after 'Area_Mean_Ratio(UMU)'
        median_ratio_index = df.columns.get_loc('Area_Mean_Ratio(UMU)') + 1
        df.insert(median_ratio_index, 'Area_Median_Ratio(UMU)', np.nan)

        # Insert 'Area_Median_Ratio(FB)' after 'Area_Median_Ratio(UMU)'
        df.insert(median_ratio_index + 1, 'Area_Median_Ratio(FB)', np.nan)

        # Insert 'Area_Median(Samples)' after 'Area_Median_Ratio(UMU)'
        df.insert(median_ratio_index, 'Area_Median(Samples)', np.nan)

        # Insert 'Area_Median(UMU blanks)' after 'Area_Median_Ratio(UMU)'
        df.insert(median_ratio_index + 1, 'Area_Median(UMU blanks)', np.nan)

        # Insert 'Area_Median(Field blanks)' after 'Area_Median_Ratio(UMU)'
        df.insert(median_ratio_index + 3, 'Area_Median(Field blanks)', np.nan)

        # Calculate Area_Median(Samples)
        df['Area_Median(Samples)'] = df[real_sample_cols].apply(
            lambda row: row[row > cutoff].median(), axis=1
        )

        # Calculate Area_Median(UMU blanks)
        df['Area_Median(UMU blanks)'] = df[umu_blank_cols].apply(
            lambda row: row[row > 0].median(), axis=1
        )

        # Calculate Area_Median(Field blanks)
        df['Area_Median(Field blanks)'] = df[field_blank_cols].apply(
            lambda row: row[row > 0].median(), axis=1
        )

        # Calculate ratios
        df['Area_Median_Ratio(UMU)'] = (
            df['Area_Median(Samples)'] / df['Area_Median(UMU blanks)']
        ).round(0).astype('Int64')

        df['Area_Median_Ratio(FB)'] = (
            df['Area_Median(Samples)'] / df['Area_Median(Field blanks)']
        ).round(0).astype('Int64')

        print(f"Calculated Area_Mean_Ratio(UMU), Area_Median_Ratio(UMU), and Area_Median_Ratio(FB).\n")

        # Assign priority
        for idx, row in df.iterrows():
            areas = row[real_sample_cols]
            areas_above_cutoff = areas[areas > cutoff]

            # Prioritize & rank exemptions
            if exempt.iloc[idx]:
                df.at[idx, 'Priority'] = 2
                df.at[idx, 'Ranking'] = 2
                df.at[idx, 'Reason'] = 'IS/RS/Mix 1 or 2/CL 1 or 2/Target hit/NORMAN hit'
                continue

            reason_list = []

            # Step 1: Priority assignment for samples:UMU blanks ratio
            if row['Area_Mean(UMU blanks)'] != 0:
                ratios = areas_above_cutoff / row['Area_Mean(UMU blanks)']
            else:
                ratios = pd.Series(dtype=float)  # Empty series if Area_Mean(UMU blanks) == 0
            valid_ratios = ratios[ratios.notna()]
            condition_coverage = len(valid_ratios) / N_real_samples if N_real_samples else 0
            if condition_coverage >= 0.10:
                condition_ratio = (valid_ratios >= 5).sum() / len(valid_ratios) 
                if condition_ratio >= 0.5:
                    df.at[idx, 'Priority'] = 2
                    df.at[idx, 'Priority_1'] = 2
                    reason_list.append(f"Step 1: {len(valid_ratios)} valid; {condition_ratio*100:.1f}% ≥5x → Priority 2")
                else:
                    df.at[idx, 'Priority'] = 0
                    df.at[idx, 'Priority_1'] = 0
                    reason_list.append(f"Step 1: {len(valid_ratios)} valid; only {condition_ratio*100:.1f}% ≥5x → Priority 0")
            else:
                df.at[idx, 'Priority'] = 0
                df.at[idx, 'Priority_1'] = 0
                reason_list.append(f"Step 1: {len(valid_ratios)} valid; coverage {condition_coverage*100:.1f}% <10% → Priority 0")
            
            # Step 2: DN/DF unknowns for 'Feature'
            if 'Feature' in str(row['Name']):
                DN_unknown = (areas_above_cutoff.count())
                DF_unknown = (DN_unknown / N_real_samples) * 100 if N_real_samples else np.nan
                df.at[idx, 'DN unknowns'] = DN_unknown
                df.at[idx, 'DF unknowns (%)'] = DF_unknown

                if DF_unknown < 2.5:
                    df.at[idx, 'Priority_2'] = 0
                    if pd.isna(df.at[idx, 'Priority']):
                        df.at[idx, 'Priority'] = 0
                        reason_list.append(f"Step 2: Unknown with DF={DF_unknown:.1f}% < 2.5%")
                    else:
                        reason_list.append(f"Step 2: Unknown with DF={DF_unknown:.1f}% < 2.5% (no override)")
                else:
                    df.at[idx, 'Priority_2'] = 2
                    reason_list.append(f"Step 2: Unknown with DF={DF_unknown:.1f}% ≥ 2.5%")

            # Step 3: DN/DF per country
            df.at[idx, 'Priority_3'] = 0
            if pd.isna(df.at[idx, 'Priority']):
                for country in country_codes:
                    if row[f'DF {country} (%)'] >= 25:
                        df.at[idx, 'Priority_3'] = 2
                        df.at[idx, 'Priority'] = 2
                        reason_list.append(f"Step 3: DF {country} = {row[f'DF {country} (%)']}% ≥ 25%")
                        break
                else:
                    reason_list.append("Step 3: All country DF < 25%")

            # Step 4: Set Priority=1 for all features that still have NaN in 'Priority'
            if pd.isna(df.at[idx, 'Priority']):
                df.at[idx, 'Priority'] = 1
                df.at[idx, 'Priority_4'] = 1
                reason_list.append("Step 4: Intermediate priority (no prior rules matched)")

            df.at[idx, 'Reason'] = "; ".join(reason_list)

        print(f"Final priority assignment completed.\n")

        # Create overall ranking categories based on combined priority & unknown DF
        def determine_ranking(row):
            # High priority: keep
            if row['Priority'] == 2:
                return 2
            # Medium priority: Priority == 0.5 or unknown features with DF unknowns >= 2.5%
            elif row['Priority'] == 1 or (('Feature' in str(row['Name'])) and (row.get('DF unknowns (%)', 0) >= 2.5)):        
                return 1
            # Otherwise, junk (delete)
            else:
                return 0

        df['Ranking'] = df.apply(determine_ranking, axis=1)
        df['Reason'] = df['Reason'].replace('', 'Step 4: Default fallback (no other conditions met)')

        print(f"Ranking assigned to features: 2=keep, 1=borderline, 0=delete.\n")
        print("Done with DF- & blank-based feature prioritization...", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        print(150 * "-")
    else:
        print("DF-based prioritization columns already exist and are populated — skipping calculation.")
        print(150 * "-")



######################### Replace Areas < Cutoff with -Cutoff ############################
def replace_small_areas(df, sample_cols, cutoff):
    user_input = input(
        f"Would you like to replace area values <{int(cutoff)} (including 0 & NA) with -{int(cutoff)}? \n"
        "0 = No \n"
        "1 = Yes \n"
        "Type your choice and press Enter: "
    ).strip()
    print(" ")
    
    if user_input == '0':
        print(f"Skipping replacement of area values <{int(cutoff)}.")
        print(150 * "-")
        return df

    # Create "_INIT" columns with the original area values preserved
    print("Creating '_INIT' columns with original sample areas...")
    # Find the insertion point: index of the earliest sample column in the current df
    sample_positions = [df.columns.get_loc(col) for col in sample_cols if col in df.columns]
    if not sample_positions:
        print("WARNING: No sample peak area columns found — nothing to preserve.")
    else:
        first_sample_pos = min(sample_positions)

        # Build list of (init_col_name, series) for columns that don't already exist
        to_insert = []
        for col in sample_cols:
            init_col = f"{col}_INIT"
            if init_col not in df.columns and col in df.columns:
                # store the series now (so future df.insert doesn't affect retrieval)
                to_insert.append((init_col, df[col].copy()))

        # Insert all INIT columns as a block before the first sample column
        for i, (init_col, series) in enumerate(to_insert):
            df.insert(first_sample_pos + i, init_col, series)

        print(f"Inserted {len(to_insert)} '_INIT' columns before the first sample column.")
    print(f"Original area columns preserved.\n")

    # Loop through sample area columns
    print(f"Replacing area values <{int(cutoff)} (including 0 & NA) with -{int(cutoff)}...", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

    for col in sample_cols:
        df[col] = pd.to_numeric(df[col], errors='coerce') # Convert to numeric, coercing 'NA', 'na', 'N/A', '' to NaN
        condition = (df[col] < cutoff) | (df[col].isna()) # TRUE for all values < cutoff, incl. NaN (NaN < cutoff is FALSE - handled separately)
        df.loc[condition, col] = -cutoff
    print("Area values replacement completed.", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    print(150 * "-")
    return df

# Interactive cutoff value selection
print("Selecting cutoff value to use for pseudo peak replacement... ")
# Determine whether cutoff values exist
have_low  = 'cutoff_low'  in globals() and int(cutoff_low)  is not None
have_high = 'cutoff_high' in globals() and int(cutoff_high) is not None
have_mean = 'cutoff_mean' in globals() and int(cutoff_mean) is not None

# If no cutoff values exist, force manual input
if not (have_low or have_high or have_mean):
    print("WARNING: No estimated cutoff values found. You must enter a custom cutoff value.")
    
    while True:
        user_cutoff = input("Enter custom cutoff value (positive number, e.g., 18250): ").strip()
        try:
            cutoff = float(user_cutoff)
            if cutoff > 0:
                break
        except:
            pass
        print("Invalid input. Please enter a positive numeric value.")
    
    print(f"Using custom cutoff: {int(cutoff)}")
    print(150 * "-")

# Use the function
df = replace_small_areas(df, sample_cols, cutoff)

# Round area values to integer
for col in sample_cols:
    df[col] = pd.to_numeric(df[col], errors='coerce').round(0).astype('Int64')



############### Populate R.I. lib with R.I. Values & Calculate R.I. delta ################
# Ask user whether to run DF- & blank-based feature prioritization
choice = input(
    "Would you like to populate 'R.I. lib' with R.I. values & calculate R.I. delta?\n"
    "0 = No\n"
    "1 = Yes\n"
    "Type your choice and press Enter: "
).strip()
print(" ")

if choice not in ["0", "1"]:
    print("WARNING: Invalid input — skipping R.I. processing.")
    print(150 * "-")
    run_ri = False
elif choice == "0":
    print("R.I. processing skipped by user.")
    print(150 * "-")
    run_ri = False
else:
    run_ri = True

# Skip entire section and continue script
if not run_ri:
    pass
else:
    print("Populating 'R.I. lib' with R.I. values & calculating R.I. delta...", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    # Insert 'R.I. source' column after 'R.I. delta'
    if 'R.I. source' not in df.columns:
        delta_idx = df.columns.get_loc('R.I. delta') + 1
        df.insert(delta_idx, 'R.I. source', pd.Series([np.nan] * len(df), dtype=object))

    # Fill missing 'R.I. lib' values and determine 'R.I. source'
    for idx, row in df.iterrows():
        if pd.notna(row['R.I. lib']):
            df.at[idx, 'R.I. source'] = 'Semi-Std_NP'
        elif pd.notna(row['RI_Semi-Std_NP']):
            df.at[idx, 'R.I. lib'] = row['RI_Semi-Std_NP']
            df.at[idx, 'R.I. source'] = 'Semi-Std_NP'
        elif pd.notna(row['RI_Std_NP']):
            df.at[idx, 'R.I. lib'] = row['RI_Std_NP']
            df.at[idx, 'R.I. source'] = 'Std_NP'
        elif pd.notna(row['RI_AI']):
            df.at[idx, 'R.I. lib'] = row['RI_AI']
            df.at[idx, 'R.I. source'] = 'AI'
        # If all are NaN, leave both columns as is (NaN)

    # Calculate R.I. delta after R.I. lib values are filled
    df['R.I. delta'] = df.apply(
        lambda row: row['R.I. calc'] - row['R.I. lib'] if pd.notna(row['R.I. calc']) and pd.notna(row['R.I. lib']) else np.nan,
        axis=1
    )

    print("Done populating 'R.I. lib' with R.I. values & calculating R.I. delta...", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    print(150 * "-")



##########################################################################################
######################### Print, Save & Format Processed File ############################
##########################################################################################

###################### Function to Format Processed Excel File ###########################
def apply_excel_formatting(file_path, cutoff_value):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Create fills
    fills = {
        "pale_yellow": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
        "pale_blue": PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid"),
        "pale_orange": PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid"),
        "pale_fuchsia": PatternFill(start_color="FFCCE5", end_color="FFCCE5", fill_type="solid"),
        "pale_red": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
        "light_grey": PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        "dark_grey": PatternFill(start_color="FF808080", end_color="FF808080", fill_type="solid"),
        "deep_yellow": PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid"),
        "green": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    }

    # Map of column header to fill
    col_color_map = {
        "ID": "pale_yellow",
        "Med RT1 (sec)": "pale_blue",
        "Med RT2 (sec)": "pale_blue",
        "R.I. calc": "pale_orange",
        "DN": "pale_fuchsia",
        "DF (%)": "pale_fuchsia",
        "Samples": "pale_red",
        "Base mass": "light_grey"
    }

    # Find header row and build column index
    header = [cell.value for cell in ws[1]]
    col_idx = {col: idx + 1 for idx, col in enumerate(header)}

    # Highlight full columns based on header
    for colname, fillkey in col_color_map.items():
        if colname in col_idx:
            col_letter = openpyxl.utils.get_column_letter(col_idx[colname])
            for row in range(1, ws.max_row + 1):
                ws[f"{col_letter}{row}"].fill = fills[fillkey]
    
    # Prevent Excel from converting CAS# to date/time
    if "CAS" in col_idx:
        col_letter = openpyxl.utils.get_column_letter(col_idx["CAS"])
        for row in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{row}"]
            cell.number_format = "@"

    # Headers alignment and row height
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[1].height = 60

    # Freeze top row
    ws.freeze_panes = "A2"

    # Set zoom level to 80%
    ws.sheet_view.zoomScale = 80

    # Set column width to 2.8
    width_2_8_cols = [
        "Merged features", "Concerns", "Edited", "Quality", "Duplicate", "Delete", "Source",
        "QM vs. BM", "Priority", "Ranking", "CL Sync2D", "RI Est, RI AI", "CommonName",
        "TargetHit", "NormanHit", "Priority_1", "Priority_2", "Priority_3", "Priority_4", "Samples"
    ]
    for col in width_2_8_cols:
        if col in col_idx:
            col_letter = openpyxl.utils.get_column_letter(col_idx[col])
            ws.column_dimensions[col_letter].width = 2.8

    # Set column width to 4.56
    width_4_6_cols = [
        "Quant mass", "Base mass", "Int(QM):Int(BM) (%)", "DN", "DN blanks", "DN UMU blanks",
        "DF (%)", "DF blanks (%)", "DF UMU blanks (%)", "DN CZ", "DF CZ (%)", "DN EE", "DF EE (%)",
        "DN IT", "DF IT (%)", "DN NL", "DF NL (%)", "DN PT", "DF PT (%)", "DN SI", "DF SI (%)",
        "DN UK", "DF UK (%)", "DN unknowns", "DF unknowns (%)", "Med RT2 (sec)", "Min RT2 (sec)",
        "Mean RT2 (sec)", "Max RT2 (sec)", "%RSD RT1", "%RSD RT2"
    ]
    for col in width_4_6_cols:
        if col in col_idx:
            col_letter = openpyxl.utils.get_column_letter(col_idx[col])
            ws.column_dimensions[col_letter].width = 4.6

    # Set column width to 5
    width_5_cols = ["0 count", "M.W.", "Similarity", "Reverse", "Probability (%)", "R.I. delta"]
    for col in width_5_cols:
        if col in col_idx:
            col_letter = openpyxl.utils.get_column_letter(col_idx[col])
            ws.column_dimensions[col_letter].width = 5
    
    # Set column width to 6
    width_6_cols = ["R.I. calc", "R.I. lib", "RI_Semi-Std_NP", "RI_Std_NP", "RI_AI"]
    for col in width_6_cols:
        if col in col_idx:
            col_letter = openpyxl.utils.get_column_letter(col_idx[col])
            ws.column_dimensions[col_letter].width = 6

    # Set column width to 8
    width_8_cols = ["Med RT1 (sec)", "Min RT1 (sec)", "Mean RT1 (sec)", "Max RT1 (sec)"]
    for col in width_8_cols:
        if col in col_idx:
            col_letter = openpyxl.utils.get_column_letter(col_idx[col])
            ws.column_dimensions[col_letter].width = 8

    # Set column width to 52
    width_52_cols = ["Name", "Class Sync2D", "Lib. search review", "Reason"]
    for col in width_52_cols:
        if col in col_idx:
            col_letter = openpyxl.utils.get_column_letter(col_idx[col])
            ws.column_dimensions[col_letter].width = 52

    # Auto-adjust selected columns (approximate by content length)
    auto_cols = ["ID", "Area ave.", "Formula", "CAS", "CAS_Consensus", "R.I. source", "Keep", "Classifications", "Area_Mean(UMU blanks)", "LOD"]
    # Safely round 'Area ave.' column if it exists
    if "Area ave." in df.columns:
        try:
            df["Area ave."] = pd.to_numeric(df["Area ave."], errors="coerce").round(0).astype("Int64")
        except Exception as e:
            print(f"WARNING: Could not round 'Area ave.': {e}")
            print("!"*150)
    for col in auto_cols:
        if col in col_idx:
            col_letter = openpyxl.utils.get_column_letter(col_idx[col])
            max_length = 0
            for row in ws.iter_rows(min_row=2, min_col=col_idx[col], max_col=col_idx[col]):
                for cell in row:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 100)
            ws.column_dimensions[col_letter].width = adjusted_width

    # Priority column
    if "Priority" in col_idx:
        col = openpyxl.utils.get_column_letter(col_idx["Priority"])
        ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", ColorScaleRule(
            start_type='num', start_value=0, start_color='F8696B',
            mid_type='num', mid_value=1, mid_color='FFEB84',
            end_type='num', end_value=2, end_color='63BE7B'))

    # Ranking column
    if "Ranking" in col_idx:
        col = openpyxl.utils.get_column_letter(col_idx["Ranking"])
        ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", ColorScaleRule(
            start_type='num', start_value=0, start_color='F8696B',
            mid_type='num', mid_value=1, mid_color='FFEB84',
            end_type='num', end_value=2, end_color='63BE7B'))

    # Values after “Samples”
    if "Samples" in col_idx:
        start_col = col_idx["Samples"] + 1
        for col in range(start_col, ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            header_value = ws[f"{col_letter}1"].value
            if header_value is None:
                continue
            range_str = f"{col_letter}2:{col_letter}{ws.max_row}"
            # 0 → grey
            ws.conditional_formatting.add(range_str, CellIsRule(operator='equal', formula=['0'], fill=fills["dark_grey"]))
            # -cutoff → deep yellow
            ws.conditional_formatting.add(range_str, CellIsRule(operator='equal', formula=[str(cutoff_value)], fill=fills["deep_yellow"]))
            ws.conditional_formatting.add(range_str, CellIsRule(operator='lessThan', formula=['0'], fill=fills["deep_yellow"]))
            # > 0 → green
            ws.conditional_formatting.add(range_str, CellIsRule(operator='greaterThan', formula=['0'], fill=fills["green"]))
    else:
        # Ask user to specify the column name after which sample peak area columns start
        col_name = input("Enter the name of the column after which sample peak area columns start: ").strip()

        # Validate that the column exists in df
        if col_name not in df.columns:
            raise KeyError(f"WARNING: Column '{col_name}' not found in DataFrame. Please check spelling and try again.\n")

        # Get the index of the column after which sample peak area columns start
        start_col = df.columns.get_loc(col_name) + 1
        print(f"Sample peak area columns will be taken starting from index {start_col} (after '{col_name}').")
        for col in range(start_col, ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(col)
            header_value = ws[f"{col_letter}1"].value
            if header_value is None:
                continue
            range_str = f"{col_letter}2:{col_letter}{ws.max_row}"
            # 0 → grey
            ws.conditional_formatting.add(range_str, CellIsRule(operator='equal', formula=['0'], fill=fills["dark_grey"]))
            # -cutoff → deep yellow
            ws.conditional_formatting.add(range_str, CellIsRule(operator='equal', formula=[str(cutoff_value)], fill=fills["deep_yellow"]))
            ws.conditional_formatting.add(range_str, CellIsRule(operator='lessThan', formula=['0'], fill=fills["deep_yellow"]))
            # > 0 → green
            ws.conditional_formatting.add(range_str, CellIsRule(operator='greaterThan', formula=['0'], fill=fills["green"]))

    # Save workbook
    wb.save(file_path)
    print(f"Formatting applied and saved to: {file_path}")
    print(150 * "-")



############# Save Full Processed File & Apply Excel File Formatting Function ############
if file_path:
    print(f"Writing data to the output file... {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    base, ext = os.path.splitext(file_path)
    processed_path = f"{base}_Prcssd.xlsx"
    df.to_excel(processed_path, index=False, engine='openpyxl')

    # Apply formatting
    apply_excel_formatting(processed_path, cutoff_value=-18250)

    print(f"Done applying formatting to the output file. {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    print(f"Formatted processed data file saved to: \n{processed_path}")
    print(150 * "-")
else:
    print(150 * "!")
    print("WARNING: No file selected.")
    print(150 * "!")

# QC printing
print("Out DF: ")
print(df)
print(150 * "-")

print(f"Processing start time: {start_time}")
print(f"Processing end time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
print(150 * "*")
print(150 * "*")
